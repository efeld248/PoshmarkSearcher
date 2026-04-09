#!/usr/bin/env python3
"""
Poshmark Deal Hunter
====================
Searches Poshmark for underpriced men's dress clothing from target brands,
compares against eBay median sold prices, and outputs an Excel report + email.

Setup (one-time):
  1. Enable 2-Step Verification on your Google account.
  2. Go to myaccount.google.com → Security → App passwords.
  3. Create a password for "Mail" and note the 16-char code.
  4. Set it as an environment variable before running:
       export GMAIL_APP_PASSWORD="xxxx xxxx xxxx xxxx"

Run manually:
  python3 poshmark_hunter.py

The script will:
  - Search Poshmark for each brand × category × size combination
  - Fetch eBay median sold prices as "book value"
  - Filter to listings ≥30% below book value
  - Save a dated Excel report to ~/Documents/Poshmark_Deals/
  - Email the report to efeld248@gmail.com (if GMAIL_APP_PASSWORD is set)
"""

import json
import logging
import os
import re
import smtplib
import time
import random
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import anthropic
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ──────────────────────────────────────────────────────────────────────────────

EMAIL_ADDRESS = "efeld248@gmail.com"
MIN_RATING_FOR_EMAIL = 4    # Only email listings rated >= this
MAX_POSHMARK_PAGES = 2      # Max category pages per brand/category/size
RATING_BATCH_SIZE = 20      # Listings rated per Claude API call
RATING_PACE_SECONDS = 3.2   # Min seconds between rater API calls (self-throttle)
CLAUDE_MODEL = "claude-haiku-4-5"

OUTPUT_DIR = Path(os.environ.get(
    "POSHMARK_OUTPUT_DIR",
    Path.home() / "Documents" / "Poshmark_Deals"
))

# Menswear brands (suits / shirts / pants)
MENSWEAR_BRANDS = [
    # Holy Grail tier
    "Yves Saint Laurent", "YSL", "Brioni", "Kiton",
    "Ermenegildo Zegna", "Zegna",
    # Corporate-appropriate Italian & Heritage American
    "Canali", "Corneliani", "Caruso", "Lubiam",
    "Aquascutum", "Daks", "Chester Barrie",
    "Hart Schaffner Marx", "Hickey Freeman", "Oxxford",
]

# Watch brands (mid-tier mechanical / heritage, target <$1000)
WATCH_BRANDS = [
    "Junghans",
    "Frederique Constant",
    "Hamilton",
]

# Trash brands to drop on sight in unfiltered verticals
BRAND_BLACKLIST = {
    "stafford", "van heusen", "arrow", "chaps", "claiborne", "kirkland",
    "dockers", "haggar", "perry ellis", "izod", "alfani", "kenneth cole",
    "calvin klein", "nautica", "tommy hilfiger", "michael kors", "apt. 9",
    "apt 9", "croft & barrow", "jos. a. bank", "jos a bank", "merona",
    "george", "geoffrey beene", "marc anthony", "sean john", "ecko",
    "joseph abboud", "pronto uomo",
}

# Category → search config
# slug   = Poshmark category URL slug
# sizes  = size filter values ([""] for none)
# brands = brand list to iterate
# max_price = optional Poshmark price ceiling
CATEGORY_CONFIG = {
    "Suit / Blazer": {
        "slug": "Men-Suits-Blazers",
        "sizes": ["42R"],
        "brands": MENSWEAR_BRANDS,
        "target_sizes": {"jacket": "42R"},
    },
    "Dress Shirt": {
        "slug": "Men-Dress-Shirts",
        "sizes": ["15.5", "15 1/2"],
        "brands": MENSWEAR_BRANDS,
        "target_sizes": {"collar": "15.5"},
    },
    "Dress Pants": {
        "slug": "Men-Dress-Pants",
        "sizes": ["32"],
        "brands": MENSWEAR_BRANDS,
        "target_sizes": {"waist": 32, "inseam": 29},
    },
    "Watch": {
        "slug": "Men-Accessories-Watches",
        "sizes": [""],
        "brands": WATCH_BRANDS,
        "max_price": 1000,
        "target_sizes": None,
    },
    # New vertical: full suits (no brand filter), scored on fit match
    "Suit (Full)": {
        "slug": "Men-Suits-Blazers",
        "sizes": ["42R"],
        "brands": [""],  # empty string = skip brand filter
        "target_sizes": {"jacket": "42R", "pants": "32Wx29L"},
    },
}

# ──────────────────────────────────────────────────────────────────────────────
# LOGGING
# ──────────────────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s %(message)s",
    handlers=[logging.StreamHandler()],
)
log = logging.getLogger(__name__)

# ──────────────────────────────────────────────────────────────────────────────
# HTTP
# ──────────────────────────────────────────────────────────────────────────────

_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
}

_SESSION = requests.Session()
_SESSION.headers.update(_HEADERS)


def _sleep(lo: float = 1.5, hi: float = 3.5) -> None:
    time.sleep(random.uniform(lo, hi))


# ──────────────────────────────────────────────────────────────────────────────
# POSHMARK SCRAPING
# ──────────────────────────────────────────────────────────────────────────────

def _extract_hex_id(slug: str) -> str | None:
    """
    Extract the 24-char MongoDB ObjectId from a Poshmark listing slug.
    e.g. 'Brioni-Silk-Tie-...-698f3721f51a0f7b83c1d2b9' → '698f3721f51a0f7b83c1d2b9'
    """
    parts = slug.rstrip("/").split("/")[-1].split("-")
    for part in reversed(parts):
        if len(part) == 24 and re.fullmatch(r"[0-9a-f]+", part):
            return part
    return None


def _extract_schema_listing_ids(html_text: str) -> list[str]:
    """Parse Schema.org ItemList JSON-LD from a Poshmark category page."""
    soup = BeautifulSoup(html_text, "html.parser")
    ids: list[str] = []
    for sc in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(sc.string or "")
            items = data.get("itemListElement") or []
            if data.get("@type") == "ItemList" and isinstance(items, list):
                for item in items:
                    u = item.get("url", "")
                    if "/listing/" in u:
                        hex_id = _extract_hex_id(u)
                        if hex_id:
                            ids.append(hex_id)
        except (json.JSONDecodeError, AttributeError):
            pass
    return ids


def _fetch_listing_detail(listing_id: str) -> dict | None:
    """Fetch full listing data from Poshmark's vm-rest API."""
    url = f"https://poshmark.com/vm-rest/posts/{listing_id}"
    try:
        _sleep(0.5, 1.5)
        resp = _SESSION.get(url, timeout=15)
        resp.raise_for_status()
        j = resp.json()
        # API returns listing data at the top level, with optional "data" wrapper
        data = j.get("data") or j
        if not isinstance(data, dict) or not data.get("id"):
            return None
        return data
    except (requests.RequestException, ValueError):
        return None


def _normalise_detail(raw: dict, category: str, fallback_brand: str) -> dict | None:
    """Normalise a vm-rest post dict into our schema."""
    try:
        pd_ = raw.get("price_amount") or {}
        price = float(pd_.get("val") or raw.get("price") or 0)
        if price <= 0:
            return None

        lid = str(raw.get("id") or "")
        if not lid:
            return None

        size_obj = raw.get("size_obj") or {}
        size_raw = size_obj.get("display") or raw.get("size") or ""

        # Original / MSRP price for price-drop detection
        orig_pd = raw.get("original_price_amount") or {}
        original_price = float(orig_pd.get("val") or raw.get("original_price") or 0)

        # Listing creation timestamp
        created_at_str = str(raw.get("created_at") or "")
        listing_age_days: int | None = None
        if created_at_str:
            try:
                created_dt = datetime.fromisoformat(
                    created_at_str.replace("Z", "+00:00")
                )
                listing_age_days = (datetime.now(created_dt.tzinfo) - created_dt).days
            except (ValueError, TypeError):
                pass

        # Price drop: current price vs original listing price
        price_drop_pct: float | None = None
        if original_price > 0 and original_price > price:
            price_drop_pct = (original_price - price) / original_price

        return {
            "id": lid,
            "title": str(raw.get("title") or ""),
            "description": str(raw.get("description") or "")[:500],
            "brand": str(raw.get("brand") or fallback_brand),
            "category": category,
            "size_raw": str(size_raw),
            "condition": str(raw.get("condition") or ""),
            "poshmark_price": price,
            "original_price": original_price if original_price > 0 else None,
            "price_drop_pct": price_drop_pct,
            "listing_age_days": listing_age_days,
            "just_listed": listing_age_days is not None and listing_age_days <= 3,
            "seller": str(raw.get("creator_username") or ""),
            "url": f"https://poshmark.com/listing/{lid}",
            "thumbnail": "",
        }
    except (TypeError, ValueError):
        return None


def _fetch_category_ids(brand: str, category: str, slug: str,
                        size: str, page: int, max_price: int | None) -> list[str]:
    """Fetch a single category-page result set and return listing IDs."""
    params = []
    if brand:
        params.append(f"brand%5B%5D={requests.utils.quote(brand)}")
    if size:
        params.append(f"size%5B%5D={requests.utils.quote(size)}")
    if max_price:
        params.append(f"max_price={max_price}")
    if page > 1:
        params.append(f"max_id={page}")
    url = f"https://poshmark.com/category/{slug}"
    if params:
        url += "?" + "&".join(params)

    try:
        resp = _SESSION.get(url, timeout=15)
        resp.raise_for_status()
        return _extract_schema_listing_ids(resp.text)
    except requests.RequestException as e:
        log.warning(f"  {brand}/{category}/{size or '-'} p{page}: {e}")
        return []


def gather_listings() -> list[dict]:
    """
    Parallel-fetch all category pages, then parallel-fetch listing details.
    Returns deduped, normalised listings across all configured brands/categories.
    """
    # 1. Build the (brand, category, slug, size, page, max_price) task list
    tasks = []
    category_targets: dict[str, dict | None] = {}
    for cat_label, cfg in CATEGORY_CONFIG.items():
        max_price = cfg.get("max_price")
        category_targets[cat_label] = cfg.get("target_sizes")
        for brand in cfg["brands"]:
            for size in cfg["sizes"]:
                for page in range(1, MAX_POSHMARK_PAGES + 1):
                    tasks.append((brand, cat_label, cfg["slug"], size, page, max_price))

    # 2. Parallel category page fetches
    # lid → (brand, category, target_sizes)
    seen: dict[str, tuple[str, str, dict | None]] = {}
    log.info(f"Fetching {len(tasks)} category pages in parallel...")
    with ThreadPoolExecutor(max_workers=12) as ex:
        futures = {
            ex.submit(_fetch_category_ids, *t): t for t in tasks
        }
        for fut in as_completed(futures):
            brand, cat_label, *_ = futures[fut]
            try:
                ids = fut.result()
            except Exception:
                continue
            for lid in ids:
                if lid not in seen:
                    seen[lid] = (brand, cat_label, category_targets.get(cat_label))
    log.info(f"  Found {len(seen)} unique listing IDs")

    # 3. Parallel listing detail fetches
    log.info(f"Fetching {len(seen)} listing details in parallel...")
    listings: list[dict] = []
    with ThreadPoolExecutor(max_workers=25) as ex:
        futures = {ex.submit(_fetch_listing_detail, lid): lid for lid in seen}
        for fut in as_completed(futures):
            lid = futures[fut]
            try:
                raw = fut.result()
            except Exception:
                continue
            if not raw:
                continue
            brand, cat_label, targets = seen[lid]
            norm = _normalise_detail(raw, cat_label, brand)
            if norm:
                norm["target_sizes"] = targets
                listings.append(norm)
    log.info(f"  Got {len(listings)} normalised listings")
    return listings


# ──────────────────────────────────────────────────────────────────────────────
# PRE-FILTER (drop obvious noise before spending API tokens)
# ──────────────────────────────────────────────────────────────────────────────

_SUIT_FULL_NEG = re.compile(
    r"\b(blazer only|jacket only|sport coat|sportcoat|top only)\b", re.I)
_SUIT_FULL_POS = re.compile(r"\bsuit(s|ing)?\b", re.I)


def _prefilter_listings(listings: list[dict]) -> list[dict]:
    """
    Apply cheap heuristic filters to drop obvious noise before rating.
    Returns the kept listings; logs a per-category count.
    """
    kept: list[dict] = []
    dropped_by_cat: dict[str, int] = {}

    for L in listings:
        cat = L.get("category", "")
        brand = (L.get("brand") or "").strip().lower()
        title = (L.get("title") or "")
        reason = None

        # Universal brand blacklist
        if brand in BRAND_BLACKLIST:
            reason = "blacklisted brand"

        # Suit (Full) vertical: must mention "suit" in title and not be an
        # explicit blazer/jacket-only listing
        elif cat == "Suit (Full)":
            if not _SUIT_FULL_POS.search(title):
                reason = "no 'suit' in title"
            elif _SUIT_FULL_NEG.search(title):
                reason = "explicit blazer/jacket-only"

        if reason:
            dropped_by_cat[cat] = dropped_by_cat.get(cat, 0) + 1
            continue
        kept.append(L)

    total_dropped = sum(dropped_by_cat.values())
    log.info(f"Pre-filter: kept {len(kept)} / {len(listings)} "
             f"(dropped {total_dropped})")
    for cat, n in sorted(dropped_by_cat.items()):
        log.info(f"  dropped {n} from {cat}")
    return kept


# ──────────────────────────────────────────────────────────────────────────────
# CLAUDE BARGAIN RATER
# ──────────────────────────────────────────────────────────────────────────────

_RATING_SYSTEM = """You are an expert appraiser of pre-owned luxury menswear and \
mid-tier mechanical watches on Poshmark. You rate each listing 1-5 on the \
INTRINSIC QUALITY AND DESIRABILITY of the item itself — NOT on price.

IGNORE the listing price entirely. Sellers inflate "original price" for marketing \
and Poshmark prices are negotiable; price tells you nothing about quality. Focus \
purely on what the item IS.

Scale:
  5 = Exceptional piece — top-tier maker, premium model/materials, excellent condition
  4 = High-quality piece — strong maker, good materials, sound condition
  3 = Solid middle-of-the-road piece
  2 = Mediocre — entry-level line, worn condition, or unremarkable
  1 = Skip — damaged, counterfeit-suspicious, mislabeled, or low-end

Judge each listing on:
  - PRODUCER PRESTIGE: Within the brand, which line/tier? (e.g. Brioni bespoke > Brioni \
RTW; Zegna Couture > Zegna Sartoria > Z Zegna; Hamilton Khaki Field/Jazzmaster > \
fashion quartz). Holy Grail names (Brioni, Kiton, Oxxford, Junghans Max Bill, \
Frederique Constant Manufacture) push toward 5.
  - MATERIALS & CONSTRUCTION: 100% wool/silk/cashmere, canvassed construction, \
Italian/British provenance, mechanical/automatic movements, in-house calibers → up. \
Polyester blends, fused construction, quartz fashion watches → down.
  - CONDITION: NWT/Excellent > Good > Fair. Damage, stains, missing parts → downgrade.
  - DESCRIPTION QUALITY: Detailed, honest descriptions with measurements/material/\
model name/provenance suggest a knowledgeable seller and authentic item. Vague or \
generic descriptions are red flags for misattribution or counterfeits.
  - TITLE KEYWORDS: "vintage", "100% wool/silk/cashmere", "made in Italy", specific \
model names ("Khaki Field", "Slimline Moonphase", "Max Bill") → positives. "Lot", \
"bundle", "AS-IS", missing model info → negatives.
  - WATCHES: Mechanical/automatic > quartz. Specific model references (caliber, \
reference number, complications) add credibility and quality.

Again: DO NOT factor price into your rating. A $30 item and a $3000 item should be \
rated identically if they are the same intrinsic quality.

In addition to the quality rating, for each listing also return:

  - "fit": 1-5 integer rating of how well this listing matches the buyer's target \
sizes (provided per listing). 5 = exact match across all target dimensions; 4 = \
matches primary dimension with minor gap on secondary; 3 = close enough to alter \
(e.g. waist matches, inseam 1-2" off); 2 = noticeable mismatch; 1 = wrong size. \
If no target_sizes are provided, return fit=0.

  - "sizing": one of "Complete", "Partial", or "Missing":
      * "Complete" = listing specifies ALL sizing dimensions relevant to the target \
(e.g. for a full suit: both jacket size AND pants waist/inseam; for a shirt: collar \
size; for pants: waist AND inseam).
      * "Partial" = some but not all dimensions specified.
      * "Missing" = no usable size info, or "see photos" dodge.
    If no target_sizes are provided, return sizing="Complete".

For the "reason" field:
  - If rating >= 3: one short phrase (max 10 words) describing what makes the item \
notable or unremarkable.
  - If rating <= 2: return reason as an empty string "". No explanation needed for \
rejects.

Respond ONLY with a JSON array, one object per listing, in the same order received:
[{"id": "...", "rating": N, "reason": "...", "fit": N, "sizing": "Complete|Partial|Missing"}]
No prose, no markdown, no code fences."""


def _load_anthropic_key() -> str:
    """Load Anthropic API key from env var or config file."""
    key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if key:
        return key
    config_file = OUTPUT_DIR / ".anthropic_api_key"
    if config_file.exists():
        return config_file.read_text().strip()
    return ""


def _build_rater_payload(batch: list[dict]) -> str:
    """Compact JSON payload for one batch. Price is intentionally omitted."""
    items = []
    for L in batch:
        items.append({
            "id": L["id"],
            "brand": L["brand"],
            "category": L["category"],
            "title": L.get("title", "")[:120],
            "condition": L.get("condition", ""),
            "size": L.get("size_raw", ""),
            "target_sizes": L.get("target_sizes"),
            "description": L.get("description", "")[:400],
        })
    return json.dumps(items, separators=(",", ":"))


def _extract_text(resp) -> str:
    """Pull text out of an Anthropic response, tolerating various block shapes."""
    try:
        for block in resp.content:
            # Object-style SDK block
            txt = getattr(block, "text", None)
            if txt:
                return txt
            # Dict-style block (rare, but seen in error paths)
            if isinstance(block, dict) and block.get("type") == "text":
                return block.get("text", "")
    except Exception:
        pass
    return ""


def _rate_batch(client: anthropic.Anthropic, batch: list[dict]) -> dict[str, dict]:
    """Send one batch of listings to Claude. Returns {id: {rating, reason, fit, sizing}}."""
    payload = _build_rater_payload(batch)
    try:
        resp = client.messages.create(
            model=CLAUDE_MODEL,
            max_tokens=2000,
            system=_RATING_SYSTEM,
            messages=[{"role": "user", "content": payload}],
        )
        text = _extract_text(resp).strip()
        if not text:
            return {}
        # Strip code fences if model added them anyway
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
        results = json.loads(text)
        return {
            r["id"]: {
                "rating": int(r.get("rating", 0)),
                "reason": str(r.get("reason", "")),
                "fit": int(r.get("fit", 0) or 0),
                "sizing": str(r.get("sizing", "")),
            }
            for r in results if isinstance(r, dict) and "id" in r
        }
    except Exception as e:
        log.warning(f"  Rating batch failed: {type(e).__name__}: {str(e)[:120]}")
        return {}


def rate_listings(listings: list[dict]) -> list[dict]:
    """
    Rate every listing 1-5 via Claude. Returns enriched listings with
    `rating` and `rating_reason` fields.
    """
    api_key = _load_anthropic_key()
    if not api_key:
        log.warning("No ANTHROPIC_API_KEY — skipping ratings.")
        for L in listings:
            L["rating"] = None
            L["rating_reason"] = ""
        return listings

    client = anthropic.Anthropic(api_key=api_key)
    batches = [listings[i:i + RATING_BATCH_SIZE]
               for i in range(0, len(listings), RATING_BATCH_SIZE)]

    log.info(f"Rating {len(listings)} listings via Claude in {len(batches)} batches "
             f"(paced {RATING_PACE_SECONDS}s apart, ETA ~{len(batches)*RATING_PACE_SECONDS/60:.1f}m)...")

    # Sequential self-throttled rater: keeps us under 10K tok/min cap without 429s
    all_ratings: dict[str, dict] = {}
    t_last = 0.0
    for i, b in enumerate(batches, 1):
        elapsed = time.time() - t_last
        if elapsed < RATING_PACE_SECONDS:
            time.sleep(RATING_PACE_SECONDS - elapsed)
        t_last = time.time()
        all_ratings.update(_rate_batch(client, b))
        if i % 10 == 0 or i == len(batches):
            log.info(f"  Batch {i}/{len(batches)}  ({len(all_ratings)} rated)")

    log.info(f"  First pass: {len(all_ratings)} / {len(listings)} rated")

    # Retry pass for any stragglers
    missing = [L for L in listings if L["id"] not in all_ratings]
    if missing:
        log.info(f"  Retrying {len(missing)} unrated listings...")
        retry_batches = [missing[i:i + RATING_BATCH_SIZE]
                         for i in range(0, len(missing), RATING_BATCH_SIZE)]
        for b in retry_batches:
            time.sleep(RATING_PACE_SECONDS)
            all_ratings.update(_rate_batch(client, b))
        log.info(f"  After retry: {len(all_ratings)} / {len(listings)} rated")

    for L in listings:
        r = all_ratings.get(L["id"], {})
        L["rating"] = r.get("rating")
        L["rating_reason"] = r.get("reason", "")
        L["fit"] = r.get("fit") or None
        L["sizing"] = r.get("sizing", "")
    return listings


# ──────────────────────────────────────────────────────────────────────────────
# EXCEL OUTPUT
# ──────────────────────────────────────────────────────────────────────────────

_COL_SPEC: list[tuple[str, int]] = [
    ("Rating",          8),
    ("Fit",             7),
    ("Sizing",         11),
    ("Brand",          18),
    ("Category",       16),
    ("Title",          42),
    ("Size",           10),
    ("Condition",      14),
    ("Price",          11),
    ("Reason",         60),
    ("Age (days)",     11),
    ("Flags",          14),
    ("Seller",         16),
    ("Link",           50),
]

_COL_COUNT = len(_COL_SPEC)


def _col(n: int) -> str:
    return get_column_letter(n)


def build_excel(deals: list[dict], output_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Deals"

    today_str = datetime.now().strftime("%B %d, %Y")

    # ── Title ────────────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{_col(_COL_COUNT)}1")
    t = ws["A1"]
    t.value = f"Poshmark Deal Hunter  ·  {today_str}"
    t.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    t.fill = PatternFill("solid", start_color="1F3864")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Headers ──────────────────────────────────────────────────────────────
    hdr_fill = PatternFill("solid", start_color="2E75B6")
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    for ci, (name, width) in enumerate(_COL_SPEC, 1):
        cell = ws.cell(row=2, column=ci, value=name)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[_col(ci)].width = width
    ws.row_dimensions[2].height = 20

    # ── Sort: highest rating first, unrated at bottom ─────────────────────────
    with_rating = sorted(
        [d for d in deals if d.get("rating") is not None],
        key=lambda x: x["rating"],
        reverse=True,
    )
    without_rating = [d for d in deals if d.get("rating") is None]
    sorted_deals = with_rating + without_rating

    # ── Data rows ─────────────────────────────────────────────────────────────
    thin = Side(style="thin", color="D0D0D0")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ri, deal in enumerate(sorted_deals, start=3):
        rating = deal.get("rating")

        # Row colour by rating
        if rating == 5:
            rfill = PatternFill("solid", start_color="C6EFCE")  # green
        elif rating == 4:
            rfill = PatternFill("solid", start_color="FFEB9C")  # amber
        elif rating == 3:
            rfill = PatternFill("solid", start_color="F2F2F2")  # grey
        else:
            rfill = PatternFill("solid", start_color="FFFFFF")

        # Build flags string
        flags = []
        if deal.get("just_listed"):
            flags.append("NEW")
        cond = deal.get("condition", "").lower()
        if "nwt" in cond or "new with" in cond:
            flags.append("NWT")
        flags_str = ", ".join(flags)

        fit_val = deal.get("fit")
        row_vals = [
            rating,
            fit_val if fit_val else "",
            deal.get("sizing", ""),
            deal.get("brand", ""),
            deal.get("category", ""),
            deal.get("title", ""),
            deal.get("size_raw", ""),
            deal.get("condition", ""),
            deal.get("poshmark_price"),
            deal.get("rating_reason", ""),
            deal.get("listing_age_days"),
            flags_str,
            deal.get("seller", ""),
            deal.get("url", ""),
        ]

        for ci, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.fill = rfill
            cell.border = bdr
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            col_name = _COL_SPEC[ci - 1][0]
            if col_name == "Price":
                cell.number_format = '"$"#,##0.00'
            elif col_name in ("Rating", "Fit"):
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name="Arial", size=12, bold=True)
            elif col_name == "Sizing":
                cell.alignment = Alignment(horizontal="center", vertical="center")
                sv = str(val).lower()
                if sv == "complete":
                    cell.fill = PatternFill("solid", start_color="C6EFCE")
                elif sv == "partial":
                    cell.fill = PatternFill("solid", start_color="FFEB9C")
                elif sv == "missing":
                    cell.fill = PatternFill("solid", start_color="FFC7CE")

    n_data = len(sorted_deals)

    # ── Freeze & filter ───────────────────────────────────────────────────────
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{_col(_COL_COUNT)}{n_data + 2}"

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    summary_rows = [
        ("Run date",              today_str),
        ("Total listings",        n_data),
        ("Rated 5 (steals)",      f"=COUNTIF(Deals!A3:A{n_data+2},5)"),
        ("Rated 4 (strong)",      f"=COUNTIF(Deals!A3:A{n_data+2},4)"),
        ("Rated 3 (fair)",        f"=COUNTIF(Deals!A3:A{n_data+2},3)"),
        ("Rated 1-2 (skip)",      f"=COUNTIF(Deals!A3:A{n_data+2},\"<=2\")"),
        ("Avg price",             f"=IFERROR(AVERAGE(Deals!I3:I{n_data+2}),\"\")"),
        ("Just listed (≤3 days)", f"=COUNTIF(Deals!L3:L{n_data+2},\"*NEW*\")"),
    ]
    for ri, (label, val) in enumerate(summary_rows, start=1):
        lc = ws2.cell(row=ri, column=1, value=label)
        vc = ws2.cell(row=ri, column=2, value=val)
        lc.font = Font(name="Arial", bold=True, size=11)
        vc.font = Font(name="Arial", size=11)
        if "price" in label.lower():
            vc.number_format = '"$"#,##0.00'
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 18

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    log.info(f"Excel saved → {output_path}")
    return output_path


# ──────────────────────────────────────────────────────────────────────────────
# EMAIL
# ──────────────────────────────────────────────────────────────────────────────

def send_email(deals: list[dict], xlsx_path: Path, app_password: str) -> None:
    top = sorted(
        [d for d in deals if (d.get("rating") or 0) >= MIN_RATING_FOR_EMAIL],
        key=lambda x: x.get("rating") or 0,
        reverse=True,
    )[:15]

    today_str = datetime.now().strftime("%B %d, %Y")

    rows_html = ""
    for d in top:
        rating = d.get("rating", 0)
        colour = "#27ae60" if rating == 5 else "#e67e22"
        stars = "★" * rating + "☆" * (5 - rating)

        rows_html += (
            f'<tr style="border-bottom:1px solid #eee;">'
            f'<td style="padding:6px 8px;text-align:center;color:{colour};font-weight:bold;font-size:14px;">{stars}</td>'
            f'<td style="padding:6px 8px;"><a href="{d["url"]}" style="color:#2980b9;text-decoration:none;">'
            f'{d["brand"]}</a></td>'
            f'<td style="padding:6px 8px;">{d["category"]}</td>'
            f'<td style="padding:6px 8px;">{d.get("title","")[:55]}</td>'
            f'<td style="padding:6px 8px;text-align:right;">${d["poshmark_price"]:.0f}</td>'
            f'<td style="padding:6px 8px;font-style:italic;color:#555;">{d.get("rating_reason","")}</td>'
            f'</tr>'
        )

    if not rows_html:
        rows_html = (
            '<tr><td colspan="6" style="padding:20px;text-align:center;color:#888;">'
            "No high-rated finds today.</td></tr>"
        )

    html_body = f"""
<html>
<body style="font-family:Arial,sans-serif;color:#222;max-width:860px;margin:0 auto;">
  <div style="background:#1F3864;padding:18px 24px;border-radius:6px 6px 0 0;">
    <h2 style="color:#fff;margin:0;">🧥 Poshmark Deal Report &mdash; {today_str}</h2>
  </div>
  <div style="padding:16px 24px;background:#f7f9fc;border:1px solid #dde3ec;border-top:none;border-radius:0 0 6px 6px;">
    <p style="margin-top:0;">
      Found <b>{len(top)}</b> listings rated {MIN_RATING_FOR_EMAIL}★ or higher by Claude.
      Full list attached as Excel.
    </p>
    <table border="0" cellspacing="0" cellpadding="0"
           style="border-collapse:collapse;width:100%;background:#fff;border:1px solid #dde3ec;border-radius:4px;">
      <thead>
        <tr style="background:#2E75B6;color:#fff;">
          <th style="padding:8px;">Rating</th>
          <th style="padding:8px;">Brand</th>
          <th style="padding:8px;">Category</th>
          <th style="padding:8px;">Title</th>
          <th style="padding:8px;">Price</th>
          <th style="padding:8px;">Why</th>
        </tr>
      </thead>
      <tbody>{rows_html}</tbody>
    </table>
    <p style="margin-bottom:0;color:#999;font-size:11px;margin-top:16px;">
      Ratings generated by Claude ({CLAUDE_MODEL}) based on brand, condition, price, and listing description.
    </p>
  </div>
</body>
</html>
"""

    msg = MIMEMultipart("mixed")
    msg["Subject"] = f"🧥 Poshmark Bargains — {today_str}  ({len(top)} top finds)"
    msg["From"] = EMAIL_ADDRESS
    msg["To"] = EMAIL_ADDRESS
    msg.attach(MIMEText(html_body, "html"))

    with open(xlsx_path, "rb") as fh:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(fh.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{xlsx_path.name}"')
    msg.attach(part)

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(EMAIL_ADDRESS, app_password)
            smtp.sendmail(EMAIL_ADDRESS, EMAIL_ADDRESS, msg.as_string())
        log.info("Email sent ✓")
    except Exception as exc:
        log.error(f"Email failed: {exc}")


# ──────────────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────────────

def _load_app_password() -> str:
    """Load Gmail app password from env var or config file."""
    pw = os.environ.get("GMAIL_APP_PASSWORD", "").strip()
    if pw:
        return pw
    config_file = OUTPUT_DIR / ".gmail_app_password"
    if config_file.exists():
        return config_file.read_text().strip()
    return ""


def main() -> Path:
    log.info("═" * 60)
    log.info("Poshmark Bargain Hunter  starting")
    log.info("═" * 60)

    listings = gather_listings()
    if not listings:
        log.warning("No listings found.")
        return OUTPUT_DIR / "empty.xlsx"

    # Pre-filter: drop obvious noise before spending API tokens
    listings = _prefilter_listings(listings)

    rated = rate_listings(listings)

    # Post-filter: Suit (Full) vertical requires Complete sizing info
    # (rater determines whether both jacket AND pant measurements are specified)
    before = len(rated)
    rated = [
        L for L in rated
        if L.get("category") != "Suit (Full)"
        or str(L.get("sizing", "")).lower() == "complete"
    ]
    dropped = before - len(rated)
    if dropped:
        log.info(f"Post-filter: dropped {dropped} Suit (Full) listings with incomplete sizing")

    log.info(f"Total listings: {len(rated)}")
    if rated:
        rating_counts = {}
        for L in rated:
            r = L.get("rating") or 0
            rating_counts[r] = rating_counts.get(r, 0) + 1
        log.info(f"  Rating distribution: {dict(sorted(rating_counts.items(), reverse=True))}")

    timestamp = datetime.now().strftime("%Y-%m-%d")
    xlsx_path = OUTPUT_DIR / f"poshmark_bargains_{timestamp}.xlsx"
    build_excel(rated, xlsx_path)

    app_pw = _load_app_password()
    if app_pw:
        send_email(rated, xlsx_path, app_pw)
    else:
        log.warning(
            "Gmail app password not configured — Excel saved but email skipped.\n"
            "  Option A: export GMAIL_APP_PASSWORD='xxxx xxxx xxxx xxxx'\n"
            f"  Option B: echo 'xxxx xxxx xxxx xxxx' > {OUTPUT_DIR / '.gmail_app_password'}"
        )

    log.info("Done.")
    return xlsx_path


if __name__ == "__main__":
    main()
