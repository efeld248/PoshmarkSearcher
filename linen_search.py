#!/usr/bin/env python3
"""
Linen Search
=============
Search-based Poshmark scraper for high-quality men's long-sleeve linen shirts.
No brand filter — rates purely on fabric, construction, colour, and condition
via two-pass Claude vision analysis.

Run:
  python3 linen_search.py
"""

import base64
import json
import logging
import os
import re
import smtplib
import time
import random
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import anthropic
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ──────────────────────────────────────────────────────────────────────────────

EMAIL_ADDRESS = "efeld248@gmail.com"
MIN_RATING_FOR_EMAIL = 4
MAX_PRICE = 150
MAX_PAGES = 10              # Pages per search query
TRIAGE_BATCH_SIZE = 20      # Text-only triage pass
VISION_BATCH_SIZE = 5       # Full vision pass
MIN_IMAGES = 3
RATING_PACE_SECONDS = 3.2
CLAUDE_MODEL = "claude-haiku-4-5"
POSH_IMG_CDN = "https://di2ponv0v5otw.cloudfront.net"

OUTPUT_DIR = Path(__file__).resolve().parent

# Multiple search queries to cast a wide net
SEARCH_QUERIES = [
    "mens linen shirt long sleeve",
    "mens linen button down shirt",
    "linen camp shirt men",
    "mens 100% linen shirt",
    "italian linen shirt men",
    "irish linen shirt men",
]
SIZES = ["M", "L"]

# ──────────────────────────────────────────────────────────────────────────────
# LOGGING
# ──────────────────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s %(message)s",
    handlers=[logging.StreamHandler()],
)
log = logging.getLogger(__name__)

# ──────────────────────────────────────────────────────────────────────────────
# HTTP
# ──────────────────────────────────────────────────────────────────────────────

_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate",
    "Connection": "keep-alive",
}

_SESSION = requests.Session()
_SESSION.headers.update(_HEADERS)
_adapter = requests.adapters.HTTPAdapter(pool_connections=20, pool_maxsize=30)
_SESSION.mount("https://", _adapter)
_SESSION.mount("http://", _adapter)


def _sleep(lo: float = 0.5, hi: float = 1.5) -> None:
    time.sleep(random.uniform(lo, hi))


# ──────────────────────────────────────────────────────────────────────────────
# POSHMARK SCRAPING (search-based)
# ──────────────────────────────────────────────────────────────────────────────

def _extract_hex_id(slug: str) -> str | None:
    parts = slug.rstrip("/").split("/")[-1].split("-")
    for part in reversed(parts):
        if len(part) == 24 and re.fullmatch(r"[0-9a-f]+", part):
            return part
    return None


def _extract_schema_listing_ids(html_text: str) -> list[str]:
    soup = BeautifulSoup(html_text, "html.parser")
    ids: list[str] = []
    for sc in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(sc.string or "")
            items = data.get("itemListElement") or []
            if data.get("@type") == "ItemList" and isinstance(items, list):
                for item in items:
                    u = item.get("url", "")
                    hid = _extract_hex_id(u)
                    if hid:
                        ids.append(hid)
        except (json.JSONDecodeError, AttributeError):
            pass
    return ids


def _fetch_search_page(query: str, size: str, page: int) -> list[str]:
    params = [
        f"query={requests.utils.quote(query)}",
        "department=Men",
        "category=Tops",
        f"size%5B%5D={requests.utils.quote(size)}",
    ]
    if MAX_PRICE:
        params.append(f"price%5B%5D=-{MAX_PRICE}")
    if page > 1:
        params.append(f"max_id={page}")
    url = f"https://poshmark.com/search?{'&'.join(params)}"
    try:
        _sleep()
        resp = _SESSION.get(url, timeout=15)
        resp.raise_for_status()
        return _extract_schema_listing_ids(resp.text)
    except Exception as e:
        log.warning(f"  search '{query}'/sz{size}/p{page}: {e}")
        return []


def _fetch_listing_detail(listing_id: str) -> dict | None:
    url = f"https://poshmark.com/vm-rest/posts/{listing_id}"
    try:
        _sleep()
        resp = _SESSION.get(url, timeout=15)
        resp.raise_for_status()
        j = resp.json()
        data = j.get("data") or j
        if not isinstance(data, dict) or not data.get("id"):
            return None
        return data
    except (requests.RequestException, ValueError):
        return None


def _normalise_detail(raw: dict) -> dict | None:
    try:
        pd_ = raw.get("price_amount") or {}
        price = float(pd_.get("val") or raw.get("price") or 0)
        if price <= 0:
            return None

        lid = str(raw.get("id") or "")
        if not lid:
            return None

        size_obj = raw.get("size_obj") or {}
        size_raw = size_obj.get("display") or raw.get("size") or ""

        created_at_str = str(raw.get("created_at") or "")
        listing_age_days: int | None = None
        if created_at_str:
            try:
                created_dt = datetime.fromisoformat(
                    created_at_str.replace("Z", "+00:00"))
                listing_age_days = (datetime.now(created_dt.tzinfo) - created_dt).days
            except (ValueError, TypeError):
                pass

        pictures = raw.get("pictures") or []
        image_urls = []
        for pic in pictures:
            path = pic.get("path_small") or pic.get("path")
            if path:
                image_urls.append(f"{POSH_IMG_CDN}/{path}")

        return {
            "id": lid,
            "title": str(raw.get("title") or ""),
            "description": str(raw.get("description") or "")[:1200],
            "brand": str(raw.get("brand") or ""),
            "size_raw": str(size_raw),
            "condition": str(raw.get("condition") or ""),
            "poshmark_price": price,
            "listing_age_days": listing_age_days,
            "just_listed": listing_age_days is not None and listing_age_days <= 3,
            "seller": str(raw.get("creator_username") or ""),
            "url": f"https://poshmark.com/listing/{lid}",
            "image_urls": image_urls,
        }
    except (TypeError, ValueError):
        return None


# ──────────────────────────────────────────────────────────────────────────────
# PRE-FILTER
# ──────────────────────────────────────────────────────────────────────────────

_LINEN_RE = re.compile(r"\b(linen|100%?\s*li)\b", re.I)
_SHORT_SLEEVE_RE = re.compile(
    r"\b(short\s+sleeve|s/s|sleeveless|tank|polo|t-?shirt)\b", re.I)


def _prefilter(listings: list[dict]) -> list[dict]:
    kept = []
    reasons: dict[str, int] = {}
    for L in listings:
        title = L.get("title", "")
        desc = L.get("description", "")
        haystack = f"{title}\n{desc}"
        reason = None

        n_imgs = len(L.get("image_urls") or [])
        if n_imgs < MIN_IMAGES:
            reason = "few images"
        elif (L.get("poshmark_price") or 0) >= MAX_PRICE:
            reason = "over price cap"
        elif not _LINEN_RE.search(haystack):
            reason = "no linen mention"
        elif _SHORT_SLEEVE_RE.search(title):
            reason = "short sleeve/polo"

        if reason:
            reasons[reason] = reasons.get(reason, 0) + 1
        else:
            kept.append(L)

    total_dropped = sum(reasons.values())
    log.info(f"Pre-filter: kept {len(kept)} / {len(kept) + total_dropped}")
    for r, n in sorted(reasons.items(), key=lambda x: -x[1]):
        log.info(f"  dropped {n}: {r}")
    return kept


def gather_listings() -> list[dict]:
    # Build task list: query × size × page
    tasks = []
    for query in SEARCH_QUERIES:
        for size in SIZES:
            for page in range(1, MAX_PAGES + 1):
                tasks.append((query, size, page))

    log.info(f"Fetching {len(tasks)} search pages in parallel...")
    seen: dict[str, None] = {}
    with ThreadPoolExecutor(max_workers=12) as ex:
        futures = {ex.submit(_fetch_search_page, q, sz, p): (q, sz, p)
                   for q, sz, p in tasks}
        for fut in as_completed(futures):
            for lid in fut.result():
                if lid not in seen:
                    seen[lid] = None

    log.info(f"  Found {len(seen)} unique listing IDs")
    log.info(f"Fetching {len(seen)} listing details in parallel...")

    listings = []
    with ThreadPoolExecutor(max_workers=25) as ex:
        futures = {ex.submit(_fetch_listing_detail, lid): lid for lid in seen}
        for fut in as_completed(futures):
            raw = fut.result()
            if raw:
                norm = _normalise_detail(raw)
                if norm:
                    listings.append(norm)

    log.info(f"  Got {len(listings)} normalised listings")
    return listings


# ──────────────────────────────────────────────────────────────────────────────
# CLAUDE RATING — TWO-PASS (triage + vision)
# ──────────────────────────────────────────────────────────────────────────────

_TRIAGE_SYSTEM = """You are a quick screener for men's linen shirts on Poshmark. \
Based ONLY on the text (brand, title, condition, description), rate 1-5:

  5 = Almost certainly a high-quality linen shirt (premium brand, detailed description, \
excellent condition, mentions 100% linen or Italian/Irish linen)
  4 = Likely good quality linen shirt
  3 = Possibly interesting — worth a closer look with photos
  2 = Probably not what we want (polyester blend, poor condition, vague description)
  1 = Definitely skip (not linen, not a shirt, damaged)

Be GENEROUS — when in doubt, rate 3. We want to avoid false negatives.

Respond ONLY with a JSON array: [{"id": "...", "rating": N}]
No prose, no markdown, no code fences."""

_RATING_SYSTEM = """You are an expert at evaluating men's linen shirts on Poshmark. \
The buyer wants long-sleeve linen shirts for summer — breathable, relaxed, versatile. \
Solid colours are preferred but not required. No brand preference — judge purely on \
the item's intrinsic quality.

Each listing includes ALL of the seller's photos. USE THE IMAGES to judge:
  - Is this actually a linen shirt? Verify fabric from labels/texture in photos.
  - Sleeve length: MUST be long-sleeve. Short sleeve, sleeveless, or polo = rating 1.
  - Colour/pattern: Solid earth tones, white, navy, sage, cream, sky blue → bonus. \
Subtle textures (linen weave, tonal stripe) are fine. Loud prints, graphics, \
Hawaiian → downgrade unless very tasteful.
  - Condition: visible stains, yellowing, pilling, holes → downgrade. \
Cross-check seller's stated condition against photos.
  - Construction quality: mother-of-pearl buttons, French seams, quality collar \
construction, camp collar or spread collar → up. Cheap plastic buttons, poor \
stitching → down.
  - Fit/style: Relaxed, slightly oversized, or tailored slim — all good for summer. \
Boxy/shapeless → slight downgrade.
  - Linen content: 100% linen or linen-dominant blend → up. \
Minor cotton or silk blend acceptable. "Linen look" polyester → rating 1.
  - Brand quality: premium fabric houses (Loro Piana, Brunello Cucinelli, Zegna, \
Turnbull & Asser, Sunspel, Todd Snyder) → bonus. Fast fashion (H&M, Zara, \
Old Navy) → neutral unless construction is surprisingly good.

IGNORE the listing price entirely.

Scale:
  5 = Exceptional — premium linen (Irish/Italian), beautiful solid colour, \
excellent condition, quality construction
  4 = Strong — good quality linen, appealing colour/pattern, sound condition
  3 = Decent — standard linen shirt, acceptable condition, nothing special
  2 = Mediocre — poor condition, unappealing pattern, or questionable linen content
  1 = Skip — not actually linen, short sleeve, damaged, or polyester blend

For each listing return:
  - "rating": 1-5 integer
  - "reason": If rating >= 3, one phrase (max 12 words) noting colour, fabric quality, \
or standout feature. If rating <= 2, return "".
  - "colour": primary colour of the shirt (e.g. "white", "navy", "sage green", \
"blue/white stripe"). Extract from photos, not just title.
  - "sleeve": "long" or "short" based on what you see in photos.
  - "linen_pct": your best estimate of linen content percentage (100, 80, 60, etc.) \
based on labels/description. If unclear, return 0.

Respond ONLY with a JSON array:
[{"id": "...", "rating": N, "reason": "...", "colour": "...", "sleeve": "...", "linen_pct": N}]
No prose, no markdown, no code fences."""


def _load_anthropic_key() -> str:
    key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if key:
        return key
    config_file = OUTPUT_DIR / ".anthropic_api_key"
    if config_file.exists():
        return config_file.read_text().strip()
    return ""


def _download_image_b64(url: str) -> str | None:
    try:
        r = _SESSION.get(url, timeout=10)
        r.raise_for_status()
        return base64.standard_b64encode(r.content).decode("ascii")
    except Exception:
        return None


def _fetch_images_for_batch(batch: list[dict]) -> dict[str, list[str]]:
    tasks = []
    for L in batch:
        for url in (L.get("image_urls") or []):
            tasks.append((L["id"], url))
    results: dict[str, list[str]] = {L["id"]: [] for L in batch}
    with ThreadPoolExecutor(max_workers=12) as ex:
        futures = {ex.submit(_download_image_b64, url): (lid, url)
                   for lid, url in tasks}
        for fut in as_completed(futures):
            lid, url = futures[fut]
            b64 = fut.result()
            if b64:
                results[lid].append(b64)
    return results


def _build_vision_content(batch: list[dict],
                          images: dict[str, list[str]]) -> list[dict]:
    content: list[dict] = []
    for L in batch:
        meta = {
            "id": L["id"],
            "brand": L["brand"],
            "title": L.get("title", "")[:120],
            "condition": L.get("condition", ""),
            "size": L.get("size_raw", ""),
            "description": L.get("description", "")[:900],
        }
        content.append({
            "type": "text",
            "text": f"=== Listing {L['id']} ===\n"
                    f"{json.dumps(meta, separators=(',', ':'))}"
        })
        for b64 in images.get(L["id"], []):
            content.append({
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": "image/jpeg",
                    "data": b64,
                }
            })
    content.append({
        "type": "text",
        "text": "Rate all listings above. Respond ONLY with the JSON array."
    })
    return content


def _extract_text(resp) -> str:
    try:
        for block in resp.content:
            txt = getattr(block, "text", None)
            if txt:
                return txt
            if isinstance(block, dict) and block.get("type") == "text":
                return block.get("text", "")
    except Exception:
        pass
    return ""


def _triage_batch(client: anthropic.Anthropic,
                  batch: list[dict]) -> dict[str, int]:
    items = [{
        "id": L["id"],
        "brand": L["brand"],
        "title": L.get("title", "")[:120],
        "condition": L.get("condition", ""),
        "description": L.get("description", "")[:400],
    } for L in batch]
    payload = json.dumps(items, separators=(",", ":"))
    try:
        resp = client.messages.create(
            model=CLAUDE_MODEL,
            max_tokens=1500,
            system=_TRIAGE_SYSTEM,
            messages=[{"role": "user", "content": payload}],
        )
        text = _extract_text(resp).strip()
        if not text:
            return {}
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
        results = json.loads(text)
        return {r["id"]: int(r.get("rating", 3))
                for r in results if isinstance(r, dict) and "id" in r}
    except Exception as e:
        log.warning(f"  Triage batch failed: {str(e)[:120]}")
        return {}


def _vision_batch(client: anthropic.Anthropic,
                  batch: list[dict]) -> dict[str, dict]:
    images = _fetch_images_for_batch(batch)
    content = _build_vision_content(batch, images)
    try:
        resp = client.messages.create(
            model=CLAUDE_MODEL,
            max_tokens=2000,
            system=_RATING_SYSTEM,
            messages=[{"role": "user", "content": content}],
        )
        text = _extract_text(resp).strip()
        if not text:
            return {}
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
        results = json.loads(text)
        return {
            r["id"]: {
                "rating": int(r.get("rating", 0)),
                "reason": str(r.get("reason", "")),
                "colour": str(r.get("colour", "")),
                "sleeve": str(r.get("sleeve", "")),
                "linen_pct": int(r.get("linen_pct", 0) or 0),
            }
            for r in results if isinstance(r, dict) and "id" in r
        }
    except Exception as e:
        log.warning(f"  Vision batch failed: {str(e)[:120]}")
        return {}


def rate_listings(listings: list[dict]) -> list[dict]:
    key = _load_anthropic_key()
    if not key:
        log.error("No Anthropic API key found")
        return listings
    client = anthropic.Anthropic(api_key=key)

    # ── Pass 1: Text-only triage ─────────────────────────────────────────
    triage_batches = [listings[i:i + TRIAGE_BATCH_SIZE]
                      for i in range(0, len(listings), TRIAGE_BATCH_SIZE)]
    log.info(f"Pass 1 (triage): {len(listings)} listings in "
             f"{len(triage_batches)} text-only batches...")

    triage_scores: dict[str, int] = {}
    t_last = 0.0
    for i, b in enumerate(triage_batches, 1):
        elapsed = time.time() - t_last
        if elapsed < RATING_PACE_SECONDS:
            time.sleep(RATING_PACE_SECONDS - elapsed)
        t_last = time.time()
        triage_scores.update(_triage_batch(client, b))
        if i % 5 == 0 or i == len(triage_batches):
            log.info(f"  Triage {i}/{len(triage_batches)}  "
                     f"({len(triage_scores)} scored)")

    TRIAGE_THRESHOLD = 3
    survivors = []
    rejects = []
    for L in listings:
        score = triage_scores.get(L["id"], 3)
        if score >= TRIAGE_THRESHOLD:
            survivors.append(L)
        else:
            L["rating"] = score
            L["rating_reason"] = ""
            L["colour"] = ""
            L["sleeve"] = ""
            L["linen_pct"] = 0
            rejects.append(L)

    log.info(f"  Triage: {len(survivors)} survivors, {len(rejects)} rejected")

    # ── Pass 2: Vision on survivors ──────────────────────────────────────
    vision_batches = [survivors[i:i + VISION_BATCH_SIZE]
                      for i in range(0, len(survivors), VISION_BATCH_SIZE)]
    log.info(f"Pass 2 (vision): {len(survivors)} listings in "
             f"{len(vision_batches)} batches "
             f"(ETA ~{len(vision_batches) * RATING_PACE_SECONDS / 60:.1f}m)...")

    all_ratings: dict[str, dict] = {}
    t_last = 0.0
    for i, b in enumerate(vision_batches, 1):
        elapsed = time.time() - t_last
        if elapsed < RATING_PACE_SECONDS:
            time.sleep(RATING_PACE_SECONDS - elapsed)
        t_last = time.time()
        all_ratings.update(_vision_batch(client, b))
        if i % 5 == 0 or i == len(vision_batches):
            log.info(f"  Vision {i}/{len(vision_batches)}  "
                     f"({len(all_ratings)} rated)")

    log.info(f"  Vision pass: {len(all_ratings)} / {len(survivors)} rated")

    # Retry stragglers
    missing = [L for L in survivors if L["id"] not in all_ratings]
    if missing:
        log.info(f"  Retrying {len(missing)} unrated...")
        for b in [missing[i:i + VISION_BATCH_SIZE]
                  for i in range(0, len(missing), VISION_BATCH_SIZE)]:
            time.sleep(RATING_PACE_SECONDS)
            all_ratings.update(_vision_batch(client, b))

    for L in survivors:
        r = all_ratings.get(L["id"], {})
        L["rating"] = r.get("rating")
        L["rating_reason"] = r.get("reason", "")
        L["colour"] = r.get("colour", "")
        L["sleeve"] = r.get("sleeve", "")
        L["linen_pct"] = r.get("linen_pct", 0)

    return rejects + survivors


# ──────────────────────────────────────────────────────────────────────────────
# EXCEL OUTPUT
# ──────────────────────────────────────────────────────────────────────────────

_COL_SPEC = [
    ("Rating",       8),
    ("Brand",       18),
    ("Title",       42),
    ("Colour",      14),
    ("Linen %",      9),
    ("Size",        10),
    ("Condition",   14),
    ("Price",       11),
    ("Offer $",     10),
    ("Reason",      50),
    ("Age (days)",  11),
    ("Flags",       14),
    ("Seller",      16),
    ("Link",        50),
]

RATING_DISCOUNT = {5: 0.82, 4: 0.75, 3: 0.65, 2: 0.50, 1: 0.40}


def _calc_offer(price, rating, age_days):
    if not price or not rating:
        return None
    base_pct = RATING_DISCOUNT.get(rating, 0.60)
    age_adj = min(age_days / 30 * 0.01, 0.10) if age_days and age_days > 0 else 0
    offer = price * (base_pct - age_adj)
    return max(round(offer / 5) * 5, 10)


def _col(n: int) -> str:
    return get_column_letter(n)


def build_excel(deals: list[dict], output_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Linen Shirts"
    n_cols = len(_COL_SPEC)

    today_str = datetime.now().strftime("%B %d, %Y")

    ws.merge_cells(f"A1:{_col(n_cols)}1")
    t = ws["A1"]
    t.value = f"Linen Search  ·  {today_str}"
    t.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    t.fill = PatternFill("solid", start_color="2D6A4F")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    hdr_fill = PatternFill("solid", start_color="40916C")
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    for ci, (name, width) in enumerate(_COL_SPEC, 1):
        cell = ws.cell(row=2, column=ci, value=name)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        ws.column_dimensions[_col(ci)].width = width
    ws.row_dimensions[2].height = 20

    with_rating = sorted(
        [d for d in deals if d.get("rating") is not None],
        key=lambda x: x["rating"], reverse=True)
    without_rating = [d for d in deals if d.get("rating") is None]
    sorted_deals = with_rating + without_rating

    thin = Side(style="thin", color="D0D0D0")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ri, deal in enumerate(sorted_deals, start=3):
        rating = deal.get("rating")
        if rating == 5:
            rfill = PatternFill("solid", start_color="D8F3DC")
        elif rating == 4:
            rfill = PatternFill("solid", start_color="F0F7DA")
        elif rating == 3:
            rfill = PatternFill("solid", start_color="F2F2F2")
        else:
            rfill = PatternFill("solid", start_color="FFFFFF")

        flags = []
        if deal.get("just_listed"):
            flags.append("NEW")
        cond = deal.get("condition", "").lower()
        if "nwt" in cond or "new with" in cond:
            flags.append("NWT")
        flags_str = ", ".join(flags)

        price = deal.get("poshmark_price")
        offer = _calc_offer(price, rating, deal.get("listing_age_days"))

        row_vals = [
            rating,
            deal.get("brand", ""),
            deal.get("title", ""),
            deal.get("colour", ""),
            deal.get("linen_pct") or "",
            deal.get("size_raw", ""),
            deal.get("condition", ""),
            price,
            offer,
            deal.get("rating_reason", ""),
            deal.get("listing_age_days"),
            flags_str,
            deal.get("seller", ""),
            deal.get("url", ""),
        ]

        for ci, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.fill = rfill
            cell.border = bdr
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            col_name = _COL_SPEC[ci - 1][0]
            if col_name in ("Price", "Offer $"):
                cell.number_format = '"$"#,##0'
                if col_name == "Offer $":
                    cell.font = Font(name="Arial", size=10, bold=True)
            elif col_name == "Rating":
                cell.alignment = Alignment(horizontal="center",
                                           vertical="center")
                cell.font = Font(name="Arial", size=12, bold=True)
            elif col_name == "Linen %":
                cell.alignment = Alignment(horizontal="center",
                                           vertical="center")

    n_data = len(sorted_deals)
    ws.freeze_panes = "A3"
    if n_data > 0:
        ws.auto_filter.ref = f"A2:{_col(n_cols)}{n_data + 2}"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    log.info(f"Excel saved → {output_path}")
    return output_path


# ──────────────────────────────────────────────────────────────────────────────
# EMAIL
# ──────────────────────────────────────────────────────────────────────────────

def send_email(deals: list[dict], xlsx_path: Path, app_password: str) -> None:
    top = sorted(
        [d for d in deals if (d.get("rating") or 0) >= MIN_RATING_FOR_EMAIL],
        key=lambda x: x.get("rating") or 0, reverse=True,
    )[:15]

    today_str = datetime.now().strftime("%B %d, %Y")
    rows_html = ""
    for d in top:
        rating = d.get("rating", 0)
        colour = "#27ae60" if rating == 5 else "#e67e22"
        stars = "★" * rating + "☆" * (5 - rating)
        rows_html += (
            f'<tr style="border-bottom:1px solid #eee;">'
            f'<td style="padding:6px 8px;text-align:center;color:{colour};'
            f'font-weight:bold;font-size:14px;">{stars}</td>'
            f'<td style="padding:6px 8px;"><a href="{d["url"]}" '
            f'style="color:#2980b9;text-decoration:none;">'
            f'{d.get("brand","")} — {d.get("title","")[:50]}</a></td>'
            f'<td style="padding:6px 8px;">{d.get("colour","")}</td>'
            f'<td style="padding:6px 8px;text-align:right;">'
            f'${d.get("poshmark_price",0):.0f}</td>'
            f'<td style="padding:6px 8px;font-size:11px;">'
            f'{d.get("rating_reason","")}</td></tr>\n'
        )

    if not rows_html:
        rows_html = ('<tr><td colspan="5" style="padding:20px;text-align:center;'
                     'color:#888;">No high-rated finds today.</td></tr>')

    html = f"""\
<html><body style="font-family:Arial,sans-serif;color:#333;">
<h2 style="color:#2D6A4F;">Linen Search — {today_str}</h2>
<p>Top {len(top)} linen finds:</p>
<table style="border-collapse:collapse;width:100%;">
<tr style="background:#2D6A4F;color:#fff;">
<th style="padding:8px;">Rating</th><th style="padding:8px;">Listing</th>
<th style="padding:8px;">Colour</th><th style="padding:8px;">Price</th>
<th style="padding:8px;">Notes</th></tr>
{rows_html}</table>
<p style="margin-top:16px;font-size:12px;color:#888;">
Full spreadsheet attached.</p>
</body></html>"""

    msg = MIMEMultipart()
    msg["From"] = EMAIL_ADDRESS
    msg["To"] = EMAIL_ADDRESS
    msg["Subject"] = f"Linen Search — {today_str}"
    msg.attach(MIMEText(html, "html"))

    part = MIMEBase("application", "octet-stream")
    with open(xlsx_path, "rb") as fh:
        part.set_payload(fh.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition",
                    f'attachment; filename="{xlsx_path.name}"')
    msg.attach(part)

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(EMAIL_ADDRESS, app_password)
        server.send_message(msg)
    log.info("Email sent ✓")


# ──────────────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────────────

def _load_app_password() -> str:
    pw = os.environ.get("GMAIL_APP_PASSWORD", "").strip()
    if pw:
        return pw
    config_file = OUTPUT_DIR / ".gmail_app_password"
    if config_file.exists():
        return config_file.read_text().strip()
    return ""


def main() -> Path:
    log.info("═" * 60)
    log.info("Linen Search  starting")
    log.info("═" * 60)

    listings = gather_listings()
    if not listings:
        log.warning("No listings found.")
        return OUTPUT_DIR / "empty.xlsx"

    listings = _prefilter(listings)
    rated = rate_listings(listings)

    # Drop short-sleeve confirmed by vision
    rated = [L for L in rated
             if L.get("sleeve", "").lower() != "short"
             and (L.get("rating") or 0) > 0]

    log.info(f"Total listings: {len(rated)}")
    if rated:
        dist = {}
        for L in rated:
            r = L.get("rating") or 0
            dist[r] = dist.get(r, 0) + 1
        log.info(f"  Rating distribution: "
                 f"{dict(sorted(dist.items(), reverse=True))}")

    today_str = datetime.now().strftime("%Y-%m-%d")
    xlsx_path = OUTPUT_DIR / f"linen_search_{today_str}.xlsx"
    build_excel(rated, xlsx_path)

    app_pw = _load_app_password()
    if app_pw:
        try:
            send_email(rated, xlsx_path, app_pw)
        except Exception as exc:
            log.error(f"Email failed: {exc}")

    log.info("Done.")
    return xlsx_path


if __name__ == "__main__":
    main()
